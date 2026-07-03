"""Export Excel formula-driven — 2 onglets : Synthèse + Comparables. Voir PROJECT_V1.md §2."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from valo.method.valuation import ValuationResult
from valo.models import TargetAnchor, ValuationRun

CLR_HEADER = "1F3864"
CLR_SECTION = "D6E4F7"
CLR_EXCL = "F2F2F2"
CLR_RESULT = "E2EFDA"
CLR_FLAG = "FCE4D6"


def _header(ws, row, col, value, bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color="FFFFFF" if bold else "000000")
    if bold:
        cell.fill = PatternFill("solid", fgColor=CLR_HEADER)
    cell.alignment = Alignment(horizontal="center")


def _fmt_m(v):
    return f"{v:.2f}x" if v is not None else "N/A"


def _growth_note(result) -> str:
    if result.price_per_pt_growth is None:
        return "croissance auto omise (données panel insuffisantes)"
    gap = (result.growth_gap or 0) * 100
    return (f"prix ≈ {result.price_per_pt_growth:.1f}x/unité × écart {gap:+.0f} pts "
            f"(médiane panel {(result.median_growth or 0) * 100:.0f} %)")


def export_excel(
    run: ValuationRun,
    anchor: TargetAnchor | None,
    result: ValuationResult,
    included_comps: list[dict],
    excluded_comps: list[dict],
    target_aggregate_value: float,
    result_ev: float,
    net_debt: float = 0.0,
    result_equity: float | None = None,
    comp_basis: str | None = None,
    flags: list[str] | None = None,
    output_dir: str = "exports",
) -> str:
    wb = Workbook()
    median_ref = _build_comps_sheet(wb, run, included_comps, excluded_comps, comp_basis or run.aggregate)
    _build_synthese_sheet(wb, run, anchor, result, target_aggregate_value, median_ref,
                          comp_basis or run.aggregate, net_debt, flags or [])
    wb.move_sheet("Synthèse", -(len(wb.sheetnames) - 1))
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    filename = f"valo_{run.target_id}_run{run.id}_{run.run_date.strftime('%Y%m%d')}.xlsx"
    path = str(Path(output_dir) / filename)
    wb.save(path)
    return path


def _build_comps_sheet(wb, run, included, excluded, agg) -> str:
    """Tableau de comps : priced en haut, proxies/exclus grisés en bas. Retourne la réf médiane."""
    ws = wb.create_sheet("Comparables")
    widths = {"A": 10, "B": 22, "C": 6, "D": 9, "E": 11, "F": 13, "G": 13, "H": 13, "I": 11, "J": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    headers = ["Ticker", "Nom", "Tier", "% CA", "Croiss. LTM", "Market Cap", "Net Debt", "EV",
               f"EV/{agg}", "Note / statut"]
    for c, h in enumerate(headers, 1):
        _header(ws, 1, c, h)

    def write_comp(row, c, greyed):
        ws.cell(row=row, column=1, value=c["ticker"])
        ws.cell(row=row, column=2, value=c["name"])
        ws.cell(row=row, column=3, value=c.get("tier"))
        pct = c.get("pct_ca_comparable")
        ws.cell(row=row, column=4, value=(pct / 100 if pct is not None else None)).number_format = "0%"
        g = c.get("revenue_growth")
        ws.cell(row=row, column=5, value=g).number_format = "0.0%"
        ws.cell(row=row, column=6, value=c.get("market_cap")).number_format = "#,##0"
        ws.cell(row=row, column=7, value=c.get("net_debt")).number_format = "#,##0"
        if greyed:
            ws.cell(row=row, column=8, value=c.get("ev")).number_format = "#,##0"
            ws.cell(row=row, column=9, value=_fmt_m(c.get("multiple")))
        else:
            ws.cell(row=row, column=8, value=f"=F{row}+G{row}").number_format = "#,##0"  # EV = mc + net debt
            agg = c.get("aggregate_value") or 0
            cell = ws.cell(row=row, column=9, value=(f"=H{row}/{agg}" if agg > 0 else "N/A"))
            cell.number_format = '0.00"x"'
        ws.cell(row=row, column=10, value=(c.get("relevance_note") or c.get("exclusion_reason") or c.get("statut") or ""))
        if greyed:
            for cc in range(1, 11):
                ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=CLR_EXCL)
                ws.cell(row=row, column=cc).font = Font(color="888888")

    row = 2
    first = row
    for c in included:
        write_comp(row, c, greyed=False)
        row += 1
    last = row - 1
    median_row = row
    ws.cell(row=median_row, column=8, value="MÉDIANE priced").font = Font(bold=True)
    med = ws.cell(row=median_row, column=9, value=f"=MEDIAN(I{first}:I{last})" if included else 0)
    med.number_format = '0.00"x"'
    med.font = Font(bold=True)
    for c in range(1, 11):
        ws.cell(row=median_row, column=c).fill = PatternFill("solid", fgColor=CLR_SECTION)
    row = median_row + 2

    if excluded:
        ws.cell(row=row, column=1, value="— Proxies / exclus (hors calcul) —").font = Font(italic=True, color="888888")
        row += 1
        for c in excluded:
            write_comp(row, c, greyed=True)
            row += 1

    return f"Comparables!I{median_row}"


def _build_synthese_sheet(wb, run, anchor, result, target_aggregate_value, median_ref: str,
                          comp_basis: str, net_debt: float, flags: list[str]):
    ws = wb.create_sheet("Synthèse")
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40

    def line(r, label, value=None, formula=None, fmt=None, note=None, bold=False, fill=None):
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=formula if formula else value)
        if fmt:
            cell.number_format = fmt
        if bold:
            cell.font = Font(bold=True)
        if note:
            ws.cell(row=r, column=3, value=note)
        if fill:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=fill)

    _header(ws, 1, 1, "Paramètre")
    _header(ws, 1, 2, "Valeur")
    _header(ws, 1, 3, "Note")

    gd = run.growth_delta or 0.0
    od = run.other_deltas or 0.0

    if result.calibrated:
        line(2, "── Ancre tour d'entrée ──", fill=CLR_SECTION, bold=True)
        line(3, "Date du tour", str(anchor.entry_date))
        line(4, f"M_entry (EV/{run.aggregate} au tour)", anchor.m_entry_aggregate, fmt='0.00"x"')
        line(5, "M_market_entry (médiane marché au tour)", anchor.m_market_entry, fmt='0.00"x"',
             note=f"basis={anchor.market_anchor_basis or '?'} · {anchor.m_market_entry_source}")
        line(7, "── Base marché ──", fill=CLR_SECTION, bold=True)
        line(8, f"Median_now priced (EV/{comp_basis})", formula=f"={median_ref}", fmt='0.00"x"',
             note=f"{result.n_priced} comps priced")
        line(9, "Moyenne winsorisée (contrôle)", result.winsor_mean, fmt='0.00"x"')
        line(10, "Dérive marché (median_now / m_market_entry)", formula="=B8/B5", fmt="0.000")
        line(11, "Base = M_entry × dérive", formula="=B4*B10", fmt='0.00"x"')
        line(13, "── Deltas société ──", fill=CLR_SECTION, bold=True)
        line(14, "Delta croissance (auto)", gd, fmt='0.00"x"', note=_growth_note(result))
        line(15, "Autres deltas (marge/NRR/taille, manuel)", od, fmt='0.00"x"')
        line(17, "── Résultat ──", fill=CLR_SECTION, bold=True)
        line(18, f"M_final (EV/{run.aggregate})", formula="=MAX(0,B11+B14+B15)", fmt='0.00"x"',
             note="max(0 ; base + deltas société)", bold=True, fill=CLR_RESULT)
        ev_row = 19
    else:
        line(2, "── Base marché (comparables directs) ──", fill=CLR_SECTION, bold=True)
        line(3, f"Median_now priced (EV/{comp_basis})", formula=f"={median_ref}", fmt='0.00"x"',
             note=f"{result.n_priced} comps priced")
        line(4, "Moyenne winsorisée (contrôle)", result.winsor_mean, fmt='0.00"x"')
        line(6, "── Deltas société ──", fill=CLR_SECTION, bold=True)
        line(7, "Delta croissance (auto)", gd, fmt='0.00"x"', note=_growth_note(result))
        line(8, "Autres deltas (marge/NRR/taille, manuel)", od, fmt='0.00"x"')
        line(10, "── Résultat ──", fill=CLR_SECTION, bold=True)
        line(11, f"M_final (EV/{run.aggregate})", formula="=MAX(0,B3+B7+B8)", fmt='0.00"x"',
             note="max(0 ; médiane + deltas société)", bold=True, fill=CLR_RESULT)
        ev_row = 12

    mfinal_ref = f"B{ev_row - 1}"
    line(ev_row, f"Agrégat cible ({run.aggregate})", target_aggregate_value, fmt="#,##0")
    line(ev_row + 1, "EV cible (100 %)", formula=f"={mfinal_ref}*B{ev_row}", fmt="#,##0",
         note="EV = M_final × agrégat cible", bold=True, fill=CLR_RESULT)
    line(ev_row + 2, "Dette nette cible", net_debt, fmt="#,##0")
    line(ev_row + 3, "Valeur des fonds propres (equity)", formula=f"=B{ev_row + 1}-B{ev_row + 2}",
         fmt="#,##0", note="Equity = EV − dette nette", bold=True, fill=CLR_RESULT)

    r = ev_row + 5
    if flags:
        line(r, "── Alertes ──", fill=CLR_FLAG, bold=True)
        r += 1
        for f in flags:
            ws.cell(row=r, column=1, value=f"⚠ {f}").font = Font(color="C55A11")
            r += 1
        r += 1
    ws.cell(row=r, column=1, value="Méthode : IPEV — médiane de dérive (set priced) + deltas société manuels.").font = Font(italic=True, color="888888")
    ws.cell(row=r + 1, column=1, value=f"Run #{run.id} · MODE {run.mode} · {run.run_date.strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, color="888888")
